import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("results.json") as f:
    data = json.load(f)

wb = openpyxl.Workbook()

title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True)
normal_font = Font(name="Calibri", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
severity_fills = {
    "Critical": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Major": PatternFill(start_color="FFEBCC", end_color="FFEBCC", fill_type="solid"),
    "Minor": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}

# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = "Summary"

ws1.merge_cells("A1:D1")
ws1["A1"] = "OBERON-301 Cross-Form Clinical Data Review - Analysis Report"
ws1["A1"].font = title_font
ws1["A1"].fill = title_fill
ws1["A1"].alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:D2")
ws1["A2"] = f"Generated: May 28, 2026 | LLM: {data['llm_provider']} ({data['llm_model']}) | Processing: {data['processing_time_seconds']}s"
ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
ws1["A2"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws1["A2"].alignment = Alignment(horizontal="center")

summary_data = [
    ("Metric", "Value"),
    ("Total Subjects Analyzed", data["total_subjects"]),
    ("Total Flags", data["total_flags"]),
    ("Critical Flags", data["critical_count"]),
    ("Major Flags", data["major_count"]),
    ("Minor Flags", data["minor_count"]),
    ("Rule-Based Flags", data["rule_flags"]),
    ("AI-Detected Flags", data["ai_flags"]),
    ("Processing Time (seconds)", data["processing_time_seconds"]),
    ("Estimated Hours Saved", data["estimated_hours_saved"]),
]

for i, (label, val) in enumerate(summary_data):
    row = i + 4
    c1 = ws1.cell(row=row, column=1, value=label)
    c2 = ws1.cell(row=row, column=2, value=val)
    c1.border = thin_border
    c2.border = thin_border
    if i == 0:
        c1.font = header_font
        c1.fill = header_fill
        c2.font = header_font
        c2.fill = header_fill
    else:
        c1.font = bold_font
        c2.font = normal_font
        if i % 2 == 0:
            c1.fill = light_gray
            c2.fill = light_gray

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20

row_offset = len(summary_data) + 6
ws1.cell(row=row_offset, column=1, value="Severity Breakdown").font = bold_font
for i, (sev, color, count) in enumerate([
    ("Critical", "FF4444", data["critical_count"]),
    ("Major", "FFA500", data["major_count"]),
    ("Minor", "44BB44", data["minor_count"]),
]):
    r = row_offset + 1 + i
    c = ws1.cell(row=r, column=1, value=sev)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    c.font = Font(bold=True, color="FFFFFF")
    ws1.cell(row=r, column=2, value=count).font = bold_font


def write_flags_sheet(ws, flags, sheet_headers, col_widths):
    for col, h in enumerate(sheet_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for i, flag in enumerate(flags):
        row = i + 2
        values = [
            i + 1,
            flag["subject_id"],
            flag["rule_id"],
            " + ".join(flag["forms_involved"]),
            flag["severity"],
            flag["source"],
            flag["confidence"],
            flag["description"],
            flag.get("suggested_query") or "",
        ]
        fill = severity_fills.get(flag["severity"])
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill and col == 5:
                cell.fill = fill
            if col == 6 and flag["source"] == "AI":
                cell.fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(sheet_headers))}{len(flags)+1}"


headers = ["#", "Subject_ID", "Rule_ID", "Forms Involved", "Severity", "Source", "Confidence", "Description", "Suggested Query"]
col_widths = [5, 12, 18, 30, 10, 8, 12, 60, 60]

# ===== Sheet 2: All Flags =====
ws2 = wb.create_sheet("All Flags")
write_flags_sheet(ws2, data["flags"], headers, col_widths)

# ===== Sheet 3: Rule Flags =====
ws3 = wb.create_sheet("Rule Flags")
rule_flags = [f for f in data["flags"] if f["source"] == "Rule"]
write_flags_sheet(ws3, rule_flags, headers, col_widths)

# ===== Sheet 4: AI Flags =====
ws4 = wb.create_sheet("AI Flags")
ai_flags = [f for f in data["flags"] if f["source"] == "AI"]
write_flags_sheet(ws4, ai_flags, headers, col_widths)

# ===== Sheet 5: By Subject =====
ws5 = wb.create_sheet("By Subject")
subjects = {}
for f in data["flags"]:
    sid = f["subject_id"]
    if sid not in subjects:
        subjects[sid] = {"critical": 0, "major": 0, "minor": 0, "rule": 0, "ai": 0, "total": 0}
    subjects[sid]["total"] += 1
    subjects[sid][f["severity"].lower()] += 1
    subjects[sid]["rule" if f["source"] == "Rule" else "ai"] += 1

subj_headers = ["Subject_ID", "Total Flags", "Critical", "Major", "Minor", "Rule", "AI"]
for col, h in enumerate(subj_headers, 1):
    cell = ws5.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, (sid, counts) in enumerate(sorted(subjects.items(), key=lambda x: -x[1]["total"])):
    row = i + 2
    vals = [sid, counts["total"], counts["critical"], counts["major"], counts["minor"], counts["rule"], counts["ai"]]
    for col, val in enumerate(vals, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if counts["critical"] > 0 and col == 3:
            cell.fill = severity_fills["Critical"]

for i, w in enumerate([12, 12, 10, 10, 10, 10, 10], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.auto_filter.ref = f"A1:G{len(subjects)+1}"

# ===== Sheet 6: Comparison =====
ws6 = wb.create_sheet("Before vs After")
ws6.merge_cells("A1:D1")
ws6["A1"] = "LLM Tuning: Before vs After Comparison"
ws6["A1"].font = title_font
ws6["A1"].fill = title_fill
ws6["A1"].alignment = Alignment(horizontal="center")

comp_headers = ["Metric", "Before Fix (Run 1)", "After Fix (Run 2)", "Target"]
for col, h in enumerate(comp_headers, 1):
    cell = ws6.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

comp_data = [
    ("Rule Engine Flags", 104, data["rule_flags"], "~104 (no change)"),
    ("AI (LLM) Flags", 94, data["ai_flags"], "10-15"),
    ("Total Flags", 198, data["total_flags"], "28-35"),
    ("Critical (Total)", 41, data["critical_count"], "~27"),
    ("Major (Total)", 156, data["major_count"], "~10"),
    ("Minor (Total)", 1, data["minor_count"], "~0"),
    ("AI Flag Reduction %", "baseline", f"{round((1 - data['ai_flags']/94)*100, 1)}%", ">90%"),
]

for i, (metric, before, after, target) in enumerate(comp_data):
    row = i + 4
    for col, val in enumerate([metric, before, after, target], 1):
        cell = ws6.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = light_gray

for c in ["A", "B", "C", "D"]:
    ws6.column_dimensions[c].width = 25

output_path = "OBERON301_Analysis_Report.xlsx"
wb.save(output_path)
print(f"Report saved: {output_path}")
print(f"  Summary | All Flags ({len(data['flags'])}) | Rule Flags ({len(rule_flags)}) | AI Flags ({len(ai_flags)}) | By Subject ({len(subjects)}) | Before vs After")
